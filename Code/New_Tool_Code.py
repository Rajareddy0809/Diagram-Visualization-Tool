import pandas as pd # type: ignore
import re
import os
import win32com.client # type: ignore
from collections import defaultdict
import warnings;    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")



from Input_paths import (
    input_file_path,
    input_folder_path,
    output_folder_path
)


table_name = "All_Tables_ID_Physical_Name"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
lookup_file_path = output_file

table_name = "Final_Unique_Edges"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
edges_input_file = output_file



table_name = "Final_Target_Input"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
targets_file = output_file


table_name = "TBMT_Boxes"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
TBMT_Input_path = output_file



### Change 1: r"Your path\バッチ一覧_v1.00.xlsx(Input File Name)"
file_path = input_file_path
xls = pd.ExcelFile(file_path)
sheet_name = "バッチ一覧"

df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
header_row_index = None
for i, row in df_raw.iterrows():
    if any("Target" in str(cell) for cell in row.values):
        header_row_index = i
        break

df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row_index)

selected_columns = {
    "Target": "Target",
    "Input": "Input Id",
    "Unnamed: 18": "Input Name",
    "Output": "Output Id",
    "Unnamed: 20": "Output Name"
}

df_selected = df[list(selected_columns.keys())]

if df_selected.iloc[0].str.contains("テーブル/ビューID|テーブル名/ビュー名", na=False).any():
    df_selected = df_selected.iloc[1:].reset_index(drop=True)

df_selected.rename(columns=selected_columns, inplace=True)

df_selected = df_selected.map(lambda x: str(x).replace("_x000D_", "").replace("\r", "") if isinstance(x, str) else x)


columns_to_select = ["Target", "Input Id", "Input Name", "Output Id", "Output Name"]
df_filtered = df_selected[columns_to_select].dropna(subset=["Input Id"]).copy()

df_filtered.loc[:, "Input Id"] = df_filtered["Input Id"].astype(str).str.split("\n")
df_filtered.loc[:, "Input Name"] = df_filtered["Input Name"].astype(str).str.split("\n")

# Clean up blank lines and whitespace in split lists
df_filtered["Input Id"] = df_filtered["Input Id"].apply(lambda lst: [x.strip() for x in lst if x.strip()])
df_filtered["Input Name"] = df_filtered["Input Name"].apply(lambda lst: [x.strip() for x in lst if x.strip()])

# Now validate based on cleaned lists
valid_rows = df_filtered["Input Id"].str.len() == df_filtered["Input Name"].str.len()
df_exploded = df_filtered[valid_rows].explode(["Input Id", "Input Name"])


valid_rows = df_filtered["Input Id"].str.len() == df_filtered["Input Name"].str.len()
df_exploded = df_filtered[valid_rows].explode(["Input Id", "Input Name"])

df_cleaned = df_exploded[
    (df_exploded["Output Name"] != "e") & 
    (~df_exploded["Input Id"].str.startswith("TBMT", na=False)) & 
    (df_exploded["Input Id"].str.match(r".*\d{5}$", na=False))
]


df_cleaned = df_cleaned.copy()
df_cleaned.loc[:, "Input-Output Pair"] = list(zip(df_cleaned["Input Id"], df_cleaned["Output Id"]))

df_final = df_cleaned.drop_duplicates(subset=["Input-Output Pair"], keep=False).drop(columns=["Input-Output Pair"])


input_ids = df_final[['Input Id', 'Input Name']].rename(columns={'Input Id': 'Unique ID', 'Input Name': 'Unique Name'})
output_ids = df_final[['Output Id', 'Output Name']].rename(columns={'Output Id': 'Unique ID', 'Output Name': 'Unique Name'})

unique_ids = pd.concat([input_ids, output_ids], axis=0).drop_duplicates().reset_index(drop=True)

def is_wk_pattern(name):
    if pd.isna(name):  
        return False
    return bool(re.search(r'WK_\d{2}$|WK\d{2}$', name))

unique_ids_filtered = unique_ids[~unique_ids['Unique Name'].apply(is_wk_pattern)]


valid_nodes_set = set(unique_ids_filtered['Unique ID'])

graph = defaultdict(list)
for _, row in df_final.iterrows():
    input_id, output_id = row['Input Id'], row['Output Id']
    graph[input_id].append(output_id)

removed_edges = set()
unique_paths = set()

def dfs(node, visited, path):
    if node in path:  
        cycle_index = path.index(node)
        removed_edges.add((path[cycle_index - 1], node))  
        return

    path.append(node)
    visited.add(node)

    for neighbor in graph.get(node, []):
        if neighbor == node:  
            removed_edges.add((node, node))  
            continue  
        dfs(neighbor, visited.copy(), path.copy())

    if len(path) > 1:
        unique_paths.add(tuple(path))  

for start_node in graph.keys():
    dfs(start_node, set(), [])

filtered_paths = set()
for path in unique_paths:
    filtered_path = tuple(node for node in path if node in valid_nodes_set)
    if len(filtered_path) > 1:  
        filtered_paths.add(filtered_path)

final_edges = set()
for path in filtered_paths:
    for i in range(len(path) - 1):
        final_edges.add((path[i], path[i + 1]))

filtered_removed_edges = set()
for edge in removed_edges:
    if edge[0] in valid_nodes_set and edge[1] in valid_nodes_set:
        final_edges.add(edge)  
        filtered_removed_edges.add(edge)  

all_nodes_in_edges = {node for edge in final_edges for node in edge if node is not None}
isolated_nodes = valid_nodes_set - all_nodes_in_edges  

for node in isolated_nodes:
    final_edges.add((node, None))  


df_final_edges = pd.DataFrame(list(final_edges), columns=['Input Id', 'Output Id'])
id_to_name = dict(zip(unique_ids_filtered["Unique ID"], unique_ids_filtered["Unique Name"]))
df_final_edges["Input Name"] = df_final_edges["Input Id"].map(id_to_name)
df_final_edges["Output Name"] = df_final_edges["Output Id"].map(id_to_name)


# Define output file paths
input_directory = os.path.dirname(file_path)
edges_output_path = os.path.join(input_directory, "Final_Unique_Edges.xlsx")
df_final_edges.to_excel(edges_output_path, index=False)



import pandas as pd
import re
import os

def clean_physical_name(name):
    """
    Remove specific prefixes from a physical name.
    """
    patterns = ["^A_", "^D_", "^D_D_", "^CDW_D_", "^M_"]
    for pattern in patterns:
        name = re.sub(pattern, "", name)
    return name.strip()

def extract_table_ids(batch_file_path, lookup_file_path, sheet_name="バッチ一覧"):
    # Load the batch file
    df = pd.read_excel(batch_file_path, sheet_name=sheet_name, skiprows=1)
    
    # Clean unwanted characters in all string cells
    df = df.map(lambda x: str(x).replace("_x000D_", "").replace("\r", "") if isinstance(x, str) else x)
    
    # Rename key columns for clarity:
    # - The first column becomes "Target"
    # - "Unnamed: 7" is assumed to be "バッチ処理物理名" (physical name)
    df.rename(columns={df.columns[0]: "Target", "Unnamed: 7": "バッチ処理物理名"}, inplace=True)
    
    # Filter rows where Target is marked with "○"
    filtered_df = df[df["Target"] == "○"].copy()
    
    # Extract unique physical names from the column (ignoring missing values)
    physical_names_raw = filtered_df["バッチ処理物理名"].dropna().unique()
    
    # Clean each physical name by removing unwanted prefixes
    cleaned_names = [clean_physical_name(name) for name in physical_names_raw]
    
    # Create a DataFrame from the cleaned names
    cleaned_df = pd.DataFrame(cleaned_names, columns=["Cleaned Physical Name"])
    
    # Load the lookup Excel file containing Table IDs and Physical Names
    lookup_df = pd.read_excel(lookup_file_path)
    if "Physical Name" not in lookup_df.columns or "Table ID" not in lookup_df.columns:
        raise ValueError("Lookup file must contain 'Physical Name' and 'Table ID' columns.")
    
    # Merge the cleaned physical names with the lookup table to get the corresponding Table IDs
    merged_df = cleaned_df.merge(lookup_df, left_on="Cleaned Physical Name", right_on="Physical Name", how="left")
    
    # Select only the columns Table ID and Physical Name for final output
    final_df = merged_df[["Table ID", "Physical Name"]].rename(columns={"Table ID": "Input ID","Physical Name": "Input Name"})
    return final_df

# === Update these file paths as needed ===
# Change 2
batch_file_path = input_file_path
lookup_file_path = lookup_file_path


# Run the extraction and mapping process
final_output = extract_table_ids(batch_file_path, lookup_file_path)

# Save the final output as an Excel file
output_file_path = os.path.join(os.path.dirname(batch_file_path), "Final_target_input.xlsx")
final_output.to_excel(output_file_path, index=False)

#print("✅ Final output saved as Final_target_input.xlsx")




import pandas as pd # type: ignore
import win32com.client # type: ignore
class Graph2D:
    def __init__(self, spacing_constant=500, vertical_spacing=120):
        self.nodes = {}  
        self.edges = {}  
        self.visited = set()  
        self.spacing_constant = spacing_constant  
        self.vertical_spacing = vertical_spacing  
        self.n = 0  
        self.y_low = 0  
        self.sink_nodes = [] 
        self.sorted_sinks = []
        self.level_first_node = {}  
        self.first_x = 0  
        self.node_depths = {}  
        self.max_depth = 0  
        self.last_y_per_level = {} 

    def add_node(self, node_id):
        if node_id not in self.nodes:
            self.nodes[node_id] = None  
            self.edges[node_id] = []  

    def add_edge(self, from_node, to_node):
        if to_node in self.edges:
            self.edges[to_node].append(from_node)
        else:
            self.edges[to_node] = [from_node]

    def find_sink_nodes(self):
        all_nodes = set(self.nodes.keys())
        nodes_with_outgoing = {parent for children in self.edges.values() for parent in children}
        return list(all_nodes - nodes_with_outgoing)

    def calculate_depths(self):
        def dfs_longest_path(node):
            if node in self.node_depths:
                return self.node_depths[node]  
            max_depth = 0  
            for parent in self.edges.get(node, []):  
                max_depth = max(max_depth, dfs_longest_path(parent) + 1)
            self.node_depths[node] = max_depth  
            return max_depth

        self.sink_nodes = self.find_sink_nodes()
        self.node_depths = {}  
        self.max_depth = 0  

        longest_path_sink = None
        max_path_length = -1
        
        
        
        for sink in self.sink_nodes:
            path_length = dfs_longest_path(sink)
            if path_length > max_path_length:
                max_path_length = path_length
                longest_path_sink = sink

        self.max_depth = max_path_length
        return longest_path_sink  

    def place_nodes(self):
        longest_path_sink = self.calculate_depths()  
        self.n = len(self.nodes)  
        self.first_x = self.n + self.max_depth * self.spacing_constant  
        self.y_low = self.n  

        if longest_path_sink:
            x = self.first_x  
            y = self.y_low  
            self.last_y_per_level[longest_path_sink] = y
            self.dfs_place(longest_path_sink, x, y, self.node_depths[longest_path_sink], is_sink=True)
        self.sorted_sinks = sorted(self.sink_nodes, key=lambda node: self.node_depths.get(node, 0), reverse=True)
        for sink in self.sorted_sinks:
            if sink != longest_path_sink and sink not in self.visited:
                x = self.first_x  
                level = self.node_depths[longest_path_sink]
                y = self.y_low + self.vertical_spacing  
                self.last_y_per_level[level] = y  
                self.y_low = y
                self.dfs_place(sink, x, y, self.node_depths[longest_path_sink], is_sink=True)

    def dfs_place(self, node, child_x, child_y, level, is_sink=False):
     if node in self.visited:
        return  

     if is_sink:
        
         x, y = child_x, child_y  
     else:
       
         x = child_x - self.spacing_constant  

     
         if level not in self.last_y_per_level:
            y = self.y_low
         else:
          
            last_y = self.last_y_per_level[level]
            y = last_y + self.vertical_spacing if last_y >= self.y_low else self.y_low

       
     self.last_y_per_level[level] = y
     self.y_low = y 

   
     self.nodes[node] = (x, y)
     self.visited.add(node)

     next_level = level - 1
     for parent in self.edges.get(node, []):  
        if parent not in self.visited:
            self.dfs_place(parent, x, y, next_level, is_sink=False)

    def get_positions(self):
        return {node: pos for node, pos in self.nodes.items() if pos is not None}

### Change 3: In the line input_file change it to =r"Your local path\Final_Unique_Edges.xlsx"  and in target_file change it to  = r"Your local path\Final_target_input.xlsx" .
input_file = edges_input_file
target_file = targets_file


df = pd.read_excel(input_file)


target_df = pd.read_excel(target_file)
if target_df.empty or "Input ID" not in target_df.columns:
    raise ValueError("❌ ERROR: No target tables found! Exiting the program.")
target_tables = target_df["Input ID"].unique().tolist()

merged_edges = set() 
unique_nodes = set() 
parent_map = {}
child_map = {}

# Mapping from IDs to names (for readability if needed)
id_to_name = {row["Input Id"]: row["Input Name"] for _, row in df.iterrows()}
id_to_name.update({row["Output Id"]: row["Output Name"] for _, row in df.iterrows()})

# Build parent and child maps
for _, row in df.iterrows():
    parent, child = row["Input Id"], row["Output Id"]
    parent_map.setdefault(child, []).append(parent)  
    child_map.setdefault(parent, []).append(child)  

# Process each target table
for target_table in target_tables:
    print(f"Processing Target Table for 1st Diagram: {target_table}...")

    # Step 1: Add Parent → Target edges (upstream)
    parents = parent_map.get(target_table, [])
    for parent in parents:
        merged_edges.add((parent, target_table))
        unique_nodes.update([parent, target_table])

    # Step 2: Traverse Children (Downstream)
    visited = set()           # to avoid re-visiting nodes
    descendants = set()

    def traverse_children(node):
        if node in visited:
            return
        visited.add(node)

        if node in child_map:
            for child in child_map[node]:
                merged_edges.add((node, child))
                unique_nodes.update([node, child])
                descendants.add(child)
                traverse_children(child)

    if target_table in child_map:
        for direct_child in child_map[target_table]:
            merged_edges.add((target_table, direct_child))
            unique_nodes.update([target_table, direct_child])
            descendants.add(direct_child)
            traverse_children(direct_child)



graph = Graph2D()
parent_map_new = {}  
child_map_new = {}  

for parent, child in merged_edges:
    
    if pd.notna(child) and child != "":
        graph.add_node(parent)
        graph.add_node(child)
        graph.add_edge(parent, child)
        parent_map_new.setdefault(child, []).append(parent) 
        child_map_new.setdefault(parent, []).append(child)  
merged_edges_df = pd.DataFrame(list(merged_edges), columns=["Parent", "Child"])


merged_edges_df["Parent Name"] = merged_edges_df["Parent"].map(id_to_name)
merged_edges_df["Child Name"] = merged_edges_df["Child"].map(id_to_name)

# ✅ **Save to Excel**

### Change 4:In the line file_path change it to = r"Your local path\Final_Unique_Edges.xlsx"
file_path = edges_input_file
input_directory = os.path.dirname(file_path)

output_file_path = os.path.join(input_directory, "Merged_Edges.xlsx")
merged_edges_df.to_excel(output_file_path, index=False)

import sys

if not merged_edges:
    print("❌ No merged edges found. Exiting the program.")
    sys.exit(1)


graph.place_nodes()
positions = graph.get_positions()







import pandas as pd # type: ignore
import os
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font, Alignment # type: ignore

def get_table_relations(target_tables, child_map, parent_map, id_to_name, output_file):
    result_data = []
    table_number = 1 
    for table in target_tables:
        descendants = set()
        visited = set()


        def traverse_children(node):
            if node in visited:
                return
            visited.add(node)
            if node in child_map:
                for child in child_map[node]:
                    if child not in descendants and child != node:
                        descendants.add(child)
                        traverse_children(child)


        if table in child_map:
            for child in child_map[table]:
                if child != table:
                    descendants.add(child)
                    traverse_children(child)

        if not descendants:
            print(f"⏭️ Skipping {table} (No descendants found)")
            continue


        one_level_parents = parent_map.get(table, [])
        table_name = id_to_name.get(table, str(table))
        
        parent_ids = [str(parent) for parent in one_level_parents if pd.notna(parent)]
        parent_names = [id_to_name.get(parent, "Unknown") for parent in parent_ids]
        child_ids = [str(child) for child in descendants if pd.notna(child)]
        child_names = [id_to_name.get(child, "Unknown") for child in child_ids]


        if not parent_ids:  
            parent_ids.append("")
            parent_names.append("")

        for i in range(max(len(parent_ids), len(child_ids), 1)):
            result_data.append({
                "Table Number": str(table_number) if i == 0 else "",
                "Target Table ID": str(table) if i == 0 else "",  
                "Target Table Name": table_name if i == 0 else "",
                "One_Level_Parent IDs": parent_ids[i] if i < len(parent_ids) else "", 
                "One_Level_Parent Names": parent_names[i] if i < len(parent_ids) else "",
                "All_Descendant IDs": child_ids[i] if i < len(child_ids) else "",
                "All_Descendant Names": child_names[i] if i < len(child_names) else "",
            })
        for _ in range(2):
             result_data.append({
                  "Target Table ID": "",
                  "Target Table Name": "",
                  "One_Level_Parent IDs": "",
                  "One_Level_Parent Names": "",
                  "All_Descendant IDs": "",
                  "All_Descendant Names": "",
                 })
        table_number += 1 

    if result_data:
        result_df = pd.DataFrame(result_data)

        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        if os.path.exists(output_file):
            os.remove(output_file)

        with pd.ExcelWriter(output_file, mode="w", engine="openpyxl") as writer:
            result_df.to_excel(writer, sheet_name="Table_Relations", index=False)


        wb = load_workbook(output_file)
        ws = wb.active


        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

      
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", wrap_text=True)


        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        
        wb.save(output_file)
        wb.close() 




output_file = os.path.join(output_folder_path, "Intermediate_3.xlsx")
output_file_new_1 = output_file


# input_directory = os.path.dirname(file_path)
# edges_output_path = os.path.join(input_directory, "Final_Unique_Edges.xlsx")
# df_final_edges.to_excel(edges_output_path, index=False)


import pandas as pd

# Step 1: Load the Excel file and parse the relevant sheet with correct header row
file_path = input_file_path # adjust path if needed
xls = pd.ExcelFile(file_path)
df = xls.parse("バッチ一覧", header=2)

# Step 2: Filter where Target is '○'
df_target = df[df['Unnamed: 0'] == '○']

# Step 3: Get unique 'バッチ処理物理名'
unique_batch_names = df_target['バッチ処理物理名'].dropna().unique()

# Step 4: Filter original data for those batch names
df_filtered_batches = df[df['バッチ処理物理名'].isin(unique_batch_names)]

# Step 5: Filter rows where 'マッピング処理物理名' ends with 'ADD_MASTER'
df_final = df_filtered_batches[
    df_filtered_batches['マッピング処理物理名'].astype(str).str.endswith('ADD_MASTER')
]

# Step 6: Build dependency mapping
dependency_rows = []

for _, row in df_final.iterrows():
    input_id = row.get('テーブル/ビューID.1')
    input_name = row.get('テーブル名/ビュー名.1')

    # These may contain multiple entries separated by newline
    dep_ids_raw = str(row.get('テーブル/ビューID'))
    dep_names_raw = str(row.get('テーブル名/ビュー名'))

    # Split by newline and clean
    dep_ids = [x.strip() for x in dep_ids_raw.split('\n') if x.strip()]
    dep_names = [x.strip() for x in dep_names_raw.split('\n') if x.strip()]

    # Match each TBMT* id and pair it with its name
    for dep_id, dep_name in zip(dep_ids, dep_names):
        if dep_id.startswith("TBMT"):
            dependency_rows.append({
                "Input ID": input_id,
                "Input Name": input_name,
                "Dependency ID": dep_id,
                "Dependency Name": dep_name
            })

# Step 7: Convert to DataFrame
df_dependencies = pd.DataFrame(dependency_rows)
input_directory = os.path.dirname(file_path)
output_file_path = os.path.join(input_directory, "TBMT_Boxes.xlsx")
df_dependencies.to_excel(output_file_path, index=False)
# Optional: Save result to Excel
# df_dependencies.to_excel("input_dependency_mapping.xlsx", index=False)





get_table_relations(target_tables, child_map_new, parent_map_new, id_to_name, output_file_new_1)




excel = win32com.client.DispatchEx("Excel.Application")

wb = excel.Workbooks.Open(output_file_new_1)
ws = wb.Sheets.Add(After=wb.Sheets(1))  
# Assuming `wb` is your workbook object
sheet_name = "Diagram"

if sheet_name in [ws.Name for ws in wb.Worksheets]:
    wb.Worksheets(sheet_name).Delete()  # Optional: delete the existing one
    # OR use a different name
    sheet_name = "Diagram_1"  # or generate a unique name dynamically

ws.Name = sheet_name

shapes = ws.Shapes
table_shapes = {}


for node, (x, y) in positions.items():
    display_name = id_to_name.get(node, node)  
    shape = shapes.AddShape(13, x, y, 160, 60)
    shape.TextFrame.Characters().Text = display_name  
    shape.Fill.ForeColor.RGB = 0xEEDD99  
    shape.TextFrame.HorizontalAlignment = 3  
    shape.Line.ForeColor.RGB = 0x000000  
    shape.Line.Weight = 1 
    shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0x000000 
    shape.TextFrame2.TextRange.ParagraphFormat.Alignment = 2  
    shape.TextFrame2.VerticalAnchor = 3 
    table_shapes[node] = shape




for child, parents in graph.edges.items():
    for parent in parents:
        if parent in table_shapes and child in table_shapes:
            shape1 = table_shapes[parent]
            shape2 = table_shapes[child]

            connector = shapes.AddConnector(2, 0, 0, 0, 0)  

            if shape1.Left > shape2.Left and shape1.Top > shape2.Top:  

              connector.ConnectorFormat.BeginConnect(shape1, 3)  
              connector.ConnectorFormat.EndConnect(shape2, 5)  

            elif shape1.Left > shape2.Left and shape1.Top < shape2.Top:  
 
              connector.ConnectorFormat.BeginConnect(shape1, 3)  
              connector.ConnectorFormat.EndConnect(shape2, 5)  

            elif shape1.Left < shape2.Left and shape1.Top > shape2.Top:  

             connector.ConnectorFormat.BeginConnect(shape1, 5)  
             connector.ConnectorFormat.EndConnect(shape2, 3)  

            elif shape1.Left < shape2.Left and shape1.Top < shape2.Top:  

             connector.ConnectorFormat.BeginConnect(shape1, 4)  
             connector.ConnectorFormat.EndConnect(shape2, 3)  

            elif shape1.Left == shape2.Left:  

             connector.ConnectorFormat.BeginConnect(shape1, 3)  
             connector.ConnectorFormat.EndConnect(shape2, 3)  

            elif shape1.Top == shape2.Top:  

                 
                 connector.ConnectorFormat.BeginConnect(shape1, 5)  
                 connector.ConnectorFormat.EndConnect(shape2, 3)  
                


            else:  

              connector.ConnectorFormat.BeginConnect(shape1, 1)  
              connector.ConnectorFormat.EndConnect(shape2, 3)  
  

            connector.Line.ForeColor.RGB = 0x000000  
            connector.Line.Weight = 2  
            connector.Line.EndArrowheadStyle = 3  





  
for sink_node in graph.sink_nodes:
    sink_shape = table_shapes.get(sink_node)  
    if sink_shape:
        sink_shape.Fill.ForeColor.RGB = 0xC1E1C1   
        sink_shape.Line.ForeColor.RGB = 0x000000  
        sink_shape.Line.Weight = 1 
        sink_shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0x000000 
        sink_shape.TextFrame2.TextRange.Font.Bold = True 




for target_table in target_tables:
    target_shape = table_shapes.get(target_table)
    if target_shape:
        target_shape.Fill.ForeColor.RGB = 0x0000FF  
        target_shape.Line.ForeColor.RGB = 0x000000  
        target_shape.Line.Weight = 1  
        target_shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0xFFFFFF  
        target_shape.TextFrame2.TextRange.Font.Bold = True  



ws.Application.ActiveWindow.DisplayGridlines = False 
ws.Application.ActiveWindow.DisplayHeadings = False   
ws.Cells.Interior.ColorIndex = 2

import pandas as pd

# Load the summary input Excel file
summary_input_file = TBMT_Input_path  # 🔁 Update path
summary_df = pd.read_excel(summary_input_file)

# Group rows by Input ID & Name
grouped = summary_df.groupby(['Input ID', 'Input Name'])

# Convert to the same structure as before
custom_summary_boxes = []
for (input_id, input_name), group in grouped:
    dependencies = list(zip(group['Dependency ID'], group['Dependency Name']))
    custom_summary_boxes.append({
        "id": input_id,
        "name": input_name,
        "dependencies": dependencies
    })


# === CONFIG ===
box_width = 250
line_height = 22
x_spacing = 300  # Space between each box

for i, box in enumerate(custom_summary_boxes):
    node_id = box["id"]
    node_name = box["name"]
    deps = box["dependencies"]

    node_pos = positions.get(node_id)
    if not node_pos:
        print(f"⚠️ Node {node_id} not found in layout. Skipping box.")
        continue

    node_x, node_y = node_pos
    base_x = node_x + x_spacing  # Position to the right of the node
    base_y = node_y

    lines = [f"{node_name} ({node_id})"]
    lines.append("--------------------")
    lines += [f"{name} ({id_})" for id_, name in deps]
    text_content = "\n".join(lines)

    shape = ws.Shapes.AddShape(1, base_x, base_y, box_width, line_height * len(lines)+1)
    shape.TextFrame.Characters().Text = text_content

    # Style the box
    shape.Fill.ForeColor.RGB = 0xDDEEFF
    shape.Line.ForeColor.RGB = 0x000000
    shape.Line.Weight = 1.5
    shape.TextFrame2.TextRange.Font.Size = 10
    shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0x000000
    shape.TextFrame2.TextRange.ParagraphFormat.Alignment = 1  # Left-align
    shape.TextFrame2.VerticalAnchor = 1  # Top-align
    shape.TextFrame.AutoSize = True
    # ✅ After creating the shape (summary box)
    original_shape = table_shapes.get(node_id)

    if original_shape:
        connector = ws.Shapes.AddConnector(2, 0, 0, 0, 0)  # Straight line

    # Connect original shape's right side ➝ to box's left side
        connector.ConnectorFormat.BeginConnect(original_shape, 5)  # Right of node
        connector.ConnectorFormat.EndConnect(shape, 2)             # Left of box

        connector.Line.ForeColor.RGB = 0x666666  # Gray line (subtle)
        connector.Line.Weight = 1.25
        connector.Line.EndArrowheadStyle = 2  # Small arrow
    else:
        print(f"⚠️ Could not find original shape for {node_id} to connect.")

    



wb.Save()
excel.Quit()




# print(f"Output saved as Final_Output_2.xlsx ")

































import pandas as pd # type: ignore
import re
import os
import win32com.client # type: ignore
from collections import defaultdict
import warnings;    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")



from Input_paths import (
    input_file_path,
    input_folder_path,
    output_folder_path
)


table_name = "All_Tables_ID_Physical_Name"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
lookup_file_path = output_file

table_name = "Final_Unique_Edges"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
edges_input_file = output_file



table_name = "Final_Target_Input"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
targets_file = output_file


table_name = "TBMT_Boxes"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
TBMT_Input_path = output_file

### Change 1: r"Your path\バッチ一覧_v1.00.xlsx(Input File Name)"
file_path = input_file_path
xls = pd.ExcelFile(file_path)
sheet_name = "バッチ一覧"

df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

header_row_index = None
for i, row in df_raw.iterrows():
    if any("Target" in str(cell) for cell in row.values):
        header_row_index = i
        break

# Step 2: Read sheet with detected header
df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row_index)

selected_columns = {
    "Target": "Target",
    "Input": "Input Id",
    "Unnamed: 18": "Input Name",
    "Output": "Output Id",
    "Unnamed: 20": "Output Name"
}
df_selected = df[list(selected_columns.keys())]

# Drop header duplication row if present
if df_selected.iloc[0].str.contains("テーブル/ビューID|テーブル名/ビュー名", na=False).any():
    df_selected = df_selected.iloc[1:].reset_index(drop=True)

df_selected.rename(columns=selected_columns, inplace=True)

# Clean encoding issues
df_selected = df_selected.map(lambda x: str(x).replace("_x000D_", "").replace("\r", "") if isinstance(x, str) else x)

# Filter and clean split input data
columns_to_select = ["Target", "Input Id", "Input Name", "Output Id", "Output Name"]
df_filtered = df_selected[columns_to_select].dropna(subset=["Input Id"]).copy()

df_filtered["Input Id"] = df_filtered["Input Id"].astype(str).str.split("\n")
df_filtered["Input Name"] = df_filtered["Input Name"].astype(str).str.split("\n")

df_filtered["Input Id"] = df_filtered["Input Id"].apply(lambda lst: [x.strip() for x in lst if x.strip()])
df_filtered["Input Name"] = df_filtered["Input Name"].apply(lambda lst: [x.strip() for x in lst if x.strip()])

valid_rows = df_filtered["Input Id"].str.len() == df_filtered["Input Name"].str.len()
df_exploded = df_filtered[valid_rows].explode(["Input Id", "Input Name"])

df_cleaned = df_exploded[
    (df_exploded["Output Name"] != "e") &
    (~df_exploded["Input Id"].str.startswith("TBMT", na=False)) &
    (df_exploded["Input Id"].str.match(r".*\d{5}$", na=False))
].copy()

df_cleaned["Input-Output Pair"] = list(zip(df_cleaned["Input Id"], df_cleaned["Output Id"]))
df_final = df_cleaned.drop_duplicates(subset=["Input-Output Pair"], keep=False).drop(columns=["Input-Output Pair"])

# Collect all unique IDs and Names
input_ids = df_final[['Input Id', 'Input Name']].rename(columns={'Input Id': 'Unique ID', 'Input Name': 'Unique Name'})
output_ids = df_final[['Output Id', 'Output Name']].rename(columns={'Output Id': 'Unique ID', 'Output Name': 'Unique Name'})
unique_ids = pd.concat([input_ids, output_ids], axis=0).drop_duplicates().reset_index(drop=True)

# Filter out WK nodes initially
def is_wk_pattern(name):
    if pd.isna(name): return False
    return bool(re.search(r'WK_\d{2}$|WK\d{2}$', name))

unique_ids_filtered = unique_ids[~unique_ids['Unique Name'].apply(is_wk_pattern)]

# STEP: Add WK nodes (properly matched from full data)

# Extract and combine Input/Output from filtered batch names
df_batch_sheet = xls.parse(sheet_name, header=2)
df_target = df_batch_sheet[df_batch_sheet['Unnamed: 0'] == '○']
unique_batch_names = df_target['バッチ処理物理名'].dropna().unique()
df_filtered_batches = df_batch_sheet[df_batch_sheet['バッチ処理物理名'].isin(unique_batch_names)]

input_id_col = 'テーブル/ビューID'
input_name_col = 'テーブル名/ビュー名'
output_id_col = 'テーブル/ビューID.1'
output_name_col = 'テーブル名/ビュー名.1'

def align_and_filter_wk_nodes(df, id_col, name_col):
    df = df[[id_col, name_col]].dropna().copy()
    df.columns = ['Unique ID', 'Unique Name']
    df['Unique ID'] = df['Unique ID'].astype(str).str.split('\n')
    df['Unique Name'] = df['Unique Name'].astype(str).str.split('\n')
    aligned_rows = []
    for _, row in df.iterrows():
        ids = [x.strip() for x in row['Unique ID'] if x.strip()]
        names = [x.strip() for x in row['Unique Name'] if x.strip()]
        if len(ids) == len(names):
            aligned_rows.extend(zip(ids, names))
    aligned_df = pd.DataFrame(aligned_rows, columns=['Unique ID', 'Unique Name'])
    return aligned_df[aligned_df['Unique Name'].apply(is_wk_pattern)].drop_duplicates().reset_index(drop=True)

wk_input_final = align_and_filter_wk_nodes(df_filtered_batches, input_id_col, input_name_col)
wk_output_final = align_and_filter_wk_nodes(df_filtered_batches, output_id_col, output_name_col)
final_clean_wk_nodes = pd.concat([wk_input_final, wk_output_final]).drop_duplicates().reset_index(drop=True)

# Merge WK nodes into valid node list
valid_nodes_set = set(unique_ids_filtered['Unique ID'])
valid_nodes_set.update(final_clean_wk_nodes["Unique ID"])
id_to_name = dict(zip(pd.concat([unique_ids_filtered, final_clean_wk_nodes])["Unique ID"],
                      pd.concat([unique_ids_filtered, final_clean_wk_nodes])["Unique Name"]))

# Build graph and detect edges
graph = defaultdict(list)
for _, row in df_final.iterrows():
    input_id, output_id = row['Input Id'], row['Output Id']
    graph[input_id].append(output_id)

removed_edges = set()
unique_paths = set()

def dfs(node, visited, path):
    if node in path:
        cycle_index = path.index(node)
        removed_edges.add((path[cycle_index - 1], node))
        return
    path.append(node)
    visited.add(node)
    for neighbor in graph.get(node, []):
        if neighbor == node:
            removed_edges.add((node, node))
            continue
        dfs(neighbor, visited.copy(), path.copy())
    if len(path) > 1:
        unique_paths.add(tuple(path))

for start_node in graph.keys():
    dfs(start_node, set(), [])

filtered_paths = set()
for path in unique_paths:
    filtered_path = tuple(node for node in path if node in valid_nodes_set)
    if len(filtered_path) > 1:
        filtered_paths.add(filtered_path)

final_edges = set()
for path in filtered_paths:
    for i in range(len(path) - 1):
        final_edges.add((path[i], path[i + 1]))

for edge in removed_edges:
    if edge[0] in valid_nodes_set and edge[1] in valid_nodes_set:
        final_edges.add(edge)

all_nodes_in_edges = {node for edge in final_edges for node in edge if node is not None}
isolated_nodes = valid_nodes_set - all_nodes_in_edges
for node in isolated_nodes:
    final_edges.add((node, None))

df_final_edges = pd.DataFrame(list(final_edges), columns=['Input Id', 'Output Id'])
df_final_edges["Input Name"] = df_final_edges["Input Id"].map(id_to_name)
df_final_edges["Output Name"] = df_final_edges["Output Id"].map(id_to_name)

# Export to Excel
input_directory = os.path.dirname(file_path)
edges_output_path = os.path.join(input_directory, "Final_Unique_Edges.xlsx")
df_final_edges.to_excel(edges_output_path, index=False)


import pandas as pd
import re
import os

def clean_physical_name(name):
    """
    Remove specific prefixes from a physical name.
    """
    patterns = ["^A_", "^D_", "^D_D_", "^CDW_D_", "^M_"]
    for pattern in patterns:
        name = re.sub(pattern, "", name)
    return name.strip()

def extract_table_ids(batch_file_path, lookup_file_path, sheet_name="バッチ一覧"):
    # Load the batch file
    df = pd.read_excel(batch_file_path, sheet_name=sheet_name, skiprows=1)
    
    # Clean unwanted characters in all string cells
    df = df.map(lambda x: str(x).replace("_x000D_", "").replace("\r", "") if isinstance(x, str) else x)
    
    # Rename key columns for clarity:
    # - The first column becomes "Target"
    # - "Unnamed: 7" is assumed to be "バッチ処理物理名" (physical name)
    df.rename(columns={df.columns[0]: "Target", "Unnamed: 7": "バッチ処理物理名"}, inplace=True)
    
    # Filter rows where Target is marked with "○"
    filtered_df = df[df["Target"] == "○"].copy()
    
    # Extract unique physical names from the column (ignoring missing values)
    physical_names_raw = filtered_df["バッチ処理物理名"].dropna().unique()
    
    # Clean each physical name by removing unwanted prefixes
    cleaned_names = [clean_physical_name(name) for name in physical_names_raw]
    
    # Create a DataFrame from the cleaned names
    cleaned_df = pd.DataFrame(cleaned_names, columns=["Cleaned Physical Name"])
    
    # Load the lookup Excel file containing Table IDs and Physical Names
    lookup_df = pd.read_excel(lookup_file_path)
    if "Physical Name" not in lookup_df.columns or "Table ID" not in lookup_df.columns:
        raise ValueError("Lookup file must contain 'Physical Name' and 'Table ID' columns.")
    
    # Merge the cleaned physical names with the lookup table to get the corresponding Table IDs
    merged_df = cleaned_df.merge(lookup_df, left_on="Cleaned Physical Name", right_on="Physical Name", how="left")
    
    # Select only the columns Table ID and Physical Name for final output
    final_df = merged_df[["Table ID", "Physical Name"]].rename(columns={"Table ID": "Input ID","Physical Name": "Input Name"})
    return final_df

# === Update these file paths as needed ===
# Change 2
batch_file_path = input_file_path
lookup_file_path = lookup_file_path


# Run the extraction and mapping process
final_output = extract_table_ids(batch_file_path, lookup_file_path)

# Save the final output as an Excel file
output_file_path = os.path.join(os.path.dirname(batch_file_path), "Final_target_input.xlsx")
final_output.to_excel(output_file_path, index=False)

#print("✅ Final output saved as Final_target_input.xlsx")







import pandas as pd # type: ignore
import re
import os
import win32com.client # type: ignore
from collections import defaultdict
import warnings;    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


from Input_paths import (
    input_file_path,
    input_folder_path,
    output_folder_path
)


table_name = "All_Tables_ID_Physical_Name"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
lookup_file_path = output_file

table_name = "Final_Unique_Edges"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
edges_input_file = output_file



table_name = "Final_Target_Input"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
targets_file = output_file


table_name = "TBMT_Boxes"  # or however you derive it
output_file = os.path.join(input_folder_path, f"{table_name}.xlsx")
TBMT_Input_path = output_file

### Change 1: r"Your path\バッチ一覧_v1.00.xlsx(Input File Name)"
file_path = input_file_path
xls = pd.ExcelFile(file_path)
sheet_name = "バッチ一覧"

df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

header_row_index = None
for i, row in df_raw.iterrows():
    if any("Target" in str(cell) for cell in row.values):
        header_row_index = i
        break

# Step 2: Read sheet with detected header
df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row_index)

selected_columns = {
    "Target": "Target",
    "Input": "Input Id",
    "Unnamed: 18": "Input Name",
    "Output": "Output Id",
    "Unnamed: 20": "Output Name"
}
df_selected = df[list(selected_columns.keys())]

# Drop header duplication row if present
if df_selected.iloc[0].str.contains("テーブル/ビューID|テーブル名/ビュー名", na=False).any():
    df_selected = df_selected.iloc[1:].reset_index(drop=True)

df_selected.rename(columns=selected_columns, inplace=True)

# Clean encoding issues
df_selected = df_selected.map(lambda x: str(x).replace("_x000D_", "").replace("\r", "") if isinstance(x, str) else x)

# Filter and clean split input data
columns_to_select = ["Target", "Input Id", "Input Name", "Output Id", "Output Name"]
df_filtered = df_selected[columns_to_select].dropna(subset=["Input Id"]).copy()

df_filtered["Input Id"] = df_filtered["Input Id"].astype(str).str.split("\n")
df_filtered["Input Name"] = df_filtered["Input Name"].astype(str).str.split("\n")

df_filtered["Input Id"] = df_filtered["Input Id"].apply(lambda lst: [x.strip() for x in lst if x.strip()])
df_filtered["Input Name"] = df_filtered["Input Name"].apply(lambda lst: [x.strip() for x in lst if x.strip()])

valid_rows = df_filtered["Input Id"].str.len() == df_filtered["Input Name"].str.len()
df_exploded = df_filtered[valid_rows].explode(["Input Id", "Input Name"])

df_cleaned = df_exploded[
    (df_exploded["Output Name"] != "e") &
    (~df_exploded["Input Id"].str.startswith("TBMT", na=False)) &
    (df_exploded["Input Id"].str.match(r".*\d{5}$", na=False))
].copy()

df_cleaned["Input-Output Pair"] = list(zip(df_cleaned["Input Id"], df_cleaned["Output Id"]))
df_final = df_cleaned.drop_duplicates(subset=["Input-Output Pair"], keep=False).drop(columns=["Input-Output Pair"])

# Collect all unique IDs and Names
input_ids = df_final[['Input Id', 'Input Name']].rename(columns={'Input Id': 'Unique ID', 'Input Name': 'Unique Name'})
output_ids = df_final[['Output Id', 'Output Name']].rename(columns={'Output Id': 'Unique ID', 'Output Name': 'Unique Name'})
unique_ids = pd.concat([input_ids, output_ids], axis=0).drop_duplicates().reset_index(drop=True)

# Filter out WK nodes initially
def is_wk_pattern(name):
    if pd.isna(name): return False
    return bool(re.search(r'WK_\d{2}$|WK\d{2}$', name))

unique_ids_filtered = unique_ids[~unique_ids['Unique Name'].apply(is_wk_pattern)]

# STEP: Add WK nodes (properly matched from full data)

# Extract and combine Input/Output from filtered batch names
df_batch_sheet = xls.parse(sheet_name, header=2)
df_target = df_batch_sheet[df_batch_sheet['Unnamed: 0'] == '○']
unique_batch_names = df_target['バッチ処理物理名'].dropna().unique()
df_filtered_batches = df_batch_sheet[df_batch_sheet['バッチ処理物理名'].isin(unique_batch_names)]

input_id_col = 'テーブル/ビューID'
input_name_col = 'テーブル名/ビュー名'
output_id_col = 'テーブル/ビューID.1'
output_name_col = 'テーブル名/ビュー名.1'

def align_and_filter_wk_nodes(df, id_col, name_col):
    df = df[[id_col, name_col]].dropna().copy()
    df.columns = ['Unique ID', 'Unique Name']
    df['Unique ID'] = df['Unique ID'].astype(str).str.split('\n')
    df['Unique Name'] = df['Unique Name'].astype(str).str.split('\n')
    aligned_rows = []
    for _, row in df.iterrows():
        ids = [x.strip() for x in row['Unique ID'] if x.strip()]
        names = [x.strip() for x in row['Unique Name'] if x.strip()]
        if len(ids) == len(names):
            aligned_rows.extend(zip(ids, names))
    aligned_df = pd.DataFrame(aligned_rows, columns=['Unique ID', 'Unique Name'])
    return aligned_df[aligned_df['Unique Name'].apply(is_wk_pattern)].drop_duplicates().reset_index(drop=True)

wk_input_final = align_and_filter_wk_nodes(df_filtered_batches, input_id_col, input_name_col)
wk_output_final = align_and_filter_wk_nodes(df_filtered_batches, output_id_col, output_name_col)
final_clean_wk_nodes = pd.concat([wk_input_final, wk_output_final]).drop_duplicates().reset_index(drop=True)

# Merge WK nodes into valid node list
valid_nodes_set = set(unique_ids_filtered['Unique ID'])
valid_nodes_set.update(final_clean_wk_nodes["Unique ID"])
id_to_name = dict(zip(pd.concat([unique_ids_filtered, final_clean_wk_nodes])["Unique ID"],
                      pd.concat([unique_ids_filtered, final_clean_wk_nodes])["Unique Name"]))

# Build graph and detect edges
graph = defaultdict(list)
for _, row in df_final.iterrows():
    input_id, output_id = row['Input Id'], row['Output Id']
    graph[input_id].append(output_id)

removed_edges = set()
unique_paths = set()

def dfs(node, visited, path):
    if node in path:
        cycle_index = path.index(node)
        removed_edges.add((path[cycle_index - 1], node))
        return
    path.append(node)
    visited.add(node)
    for neighbor in graph.get(node, []):
        if neighbor == node:
            removed_edges.add((node, node))
            continue
        dfs(neighbor, visited.copy(), path.copy())
    if len(path) > 1:
        unique_paths.add(tuple(path))

for start_node in graph.keys():
    dfs(start_node, set(), [])

filtered_paths = set()
for path in unique_paths:
    filtered_path = tuple(node for node in path if node in valid_nodes_set)
    if len(filtered_path) > 1:
        filtered_paths.add(filtered_path)

final_edges = set()
for path in filtered_paths:
    for i in range(len(path) - 1):
        final_edges.add((path[i], path[i + 1]))

for edge in removed_edges:
    if edge[0] in valid_nodes_set and edge[1] in valid_nodes_set:
        final_edges.add(edge)

all_nodes_in_edges = {node for edge in final_edges for node in edge if node is not None}
isolated_nodes = valid_nodes_set - all_nodes_in_edges
for node in isolated_nodes:
    final_edges.add((node, None))

df_final_edges = pd.DataFrame(list(final_edges), columns=['Input Id', 'Output Id'])
df_final_edges["Input Name"] = df_final_edges["Input Id"].map(id_to_name)
df_final_edges["Output Name"] = df_final_edges["Output Id"].map(id_to_name)

# Export to Excel
input_directory = os.path.dirname(file_path)
edges_output_path = os.path.join(input_directory, "Final_Unique_Edges.xlsx")
df_final_edges.to_excel(edges_output_path, index=False)


import pandas as pd
import re
import os

def clean_physical_name(name):
    """
    Remove specific prefixes from a physical name.
    """
    patterns = ["^A_", "^D_", "^D_D_", "^CDW_D_", "^M_"]
    for pattern in patterns:
        name = re.sub(pattern, "", name)
    return name.strip()

def extract_table_ids(batch_file_path, lookup_file_path, sheet_name="バッチ一覧"):
    # Load the batch file
    df = pd.read_excel(batch_file_path, sheet_name=sheet_name, skiprows=1)
    
    # Clean unwanted characters in all string cells
    df = df.map(lambda x: str(x).replace("_x000D_", "").replace("\r", "") if isinstance(x, str) else x)
    
    # Rename key columns for clarity:
    # - The first column becomes "Target"
    # - "Unnamed: 7" is assumed to be "バッチ処理物理名" (physical name)
    df.rename(columns={df.columns[0]: "Target", "Unnamed: 7": "バッチ処理物理名"}, inplace=True)
    
    # Filter rows where Target is marked with "○"
    filtered_df = df[df["Target"] == "○"].copy()
    
    # Extract unique physical names from the column (ignoring missing values)
    physical_names_raw = filtered_df["バッチ処理物理名"].dropna().unique()
    
    # Clean each physical name by removing unwanted prefixes
    cleaned_names = [clean_physical_name(name) for name in physical_names_raw]
    
    # Create a DataFrame from the cleaned names
    cleaned_df = pd.DataFrame(cleaned_names, columns=["Cleaned Physical Name"])
    
    # Load the lookup Excel file containing Table IDs and Physical Names
    lookup_df = pd.read_excel(lookup_file_path)
    if "Physical Name" not in lookup_df.columns or "Table ID" not in lookup_df.columns:
        raise ValueError("Lookup file must contain 'Physical Name' and 'Table ID' columns.")
    
    # Merge the cleaned physical names with the lookup table to get the corresponding Table IDs
    merged_df = cleaned_df.merge(lookup_df, left_on="Cleaned Physical Name", right_on="Physical Name", how="left")
    
    # Select only the columns Table ID and Physical Name for final output
    final_df = merged_df[["Table ID", "Physical Name"]].rename(columns={"Table ID": "Input ID","Physical Name": "Input Name"})
    return final_df

# === Update these file paths as needed ===
# Change 2
batch_file_path = input_file_path
lookup_file_path = lookup_file_path


# Run the extraction and mapping process
final_output = extract_table_ids(batch_file_path, lookup_file_path)

# Save the final output as an Excel file
output_file_path = os.path.join(os.path.dirname(batch_file_path), "Final_target_input.xlsx")
final_output.to_excel(output_file_path, index=False)

#print("✅ Final output saved as Final_target_input.xlsx")




import pandas as pd # type: ignore
import win32com.client # type: ignore
class Graph2D:
    def __init__(self, spacing_constant=500, vertical_spacing=120):
        self.nodes = {}  
        self.edges = {}  
        self.visited = set()  
        self.spacing_constant = spacing_constant  
        self.vertical_spacing = vertical_spacing  
        self.n = 0  
        self.y_low = 0  
        self.sink_nodes = [] 
        self.sorted_sinks = []
        self.level_first_node = {}  
        self.first_x = 0  
        self.node_depths = {}  
        self.max_depth = 0  
        self.last_y_per_level = {} 

    def add_node(self, node_id):
        if node_id not in self.nodes:
            self.nodes[node_id] = None  
            self.edges[node_id] = []  

    def add_edge(self, from_node, to_node):
        if to_node in self.edges:
            self.edges[to_node].append(from_node)
        else:
            self.edges[to_node] = [from_node]

    def find_sink_nodes(self):
        all_nodes = set(self.nodes.keys())
        nodes_with_outgoing = {parent for children in self.edges.values() for parent in children}
        return list(all_nodes - nodes_with_outgoing)

    def calculate_depths(self):
        def dfs_longest_path(node):
            if node in self.node_depths:
                return self.node_depths[node]  
            max_depth = 0  
            for parent in self.edges.get(node, []):  
                max_depth = max(max_depth, dfs_longest_path(parent) + 1)
            self.node_depths[node] = max_depth  
            return max_depth

        self.sink_nodes = self.find_sink_nodes()
        self.node_depths = {}  
        self.max_depth = 0  

        longest_path_sink = None
        max_path_length = -1
        
        
        
        for sink in self.sink_nodes:
            path_length = dfs_longest_path(sink)
            if path_length > max_path_length:
                max_path_length = path_length
                longest_path_sink = sink

        self.max_depth = max_path_length
        return longest_path_sink  

    def place_nodes(self):
        longest_path_sink = self.calculate_depths()  
        self.n = len(self.nodes)  
        self.first_x = self.n + self.max_depth * self.spacing_constant  
        self.y_low = self.n  

        if longest_path_sink:
            x = self.first_x  
            y = self.y_low  
            self.last_y_per_level[longest_path_sink] = y
            self.dfs_place(longest_path_sink, x, y, self.node_depths[longest_path_sink], is_sink=True)
        self.sorted_sinks = sorted(self.sink_nodes, key=lambda node: self.node_depths.get(node, 0), reverse=True)
        for sink in self.sorted_sinks:
            if sink != longest_path_sink and sink not in self.visited:
                x = self.first_x  
                level = self.node_depths[longest_path_sink]
                y = self.y_low + self.vertical_spacing  
                self.last_y_per_level[level] = y  
                self.y_low = y
                self.dfs_place(sink, x, y, self.node_depths[longest_path_sink], is_sink=True)

    def dfs_place(self, node, child_x, child_y, level, is_sink=False):
     if node in self.visited:
        return  

     if is_sink:
        
         x, y = child_x, child_y  
     else:
       
         x = child_x - self.spacing_constant  

     
         if level not in self.last_y_per_level:
            y = self.y_low
         else:
          
            last_y = self.last_y_per_level[level]
            y = last_y + self.vertical_spacing if last_y >= self.y_low else self.y_low

       
     self.last_y_per_level[level] = y
     self.y_low = y 

   
     self.nodes[node] = (x, y)
     self.visited.add(node)

     next_level = level - 1
     for parent in self.edges.get(node, []):  
        if parent not in self.visited:
            self.dfs_place(parent, x, y, next_level, is_sink=False)

    def get_positions(self):
        return {node: pos for node, pos in self.nodes.items() if pos is not None}

### Change 3: In the line input_file change it to =r"Your local path\Final_Unique_Edges.xlsx"  and in target_file change it to  = r"Your local path\Final_target_input.xlsx" .
input_file = edges_input_file
target_file = targets_file


df = pd.read_excel(input_file)


target_df = pd.read_excel(target_file)
if target_df.empty or "Input ID" not in target_df.columns:
    raise ValueError("❌ ERROR: No target tables found! Exiting the program.")
target_tables = target_df["Input ID"].unique().tolist()
merged_edges = set() 
unique_nodes = set() 
parent_map = {}
child_map = {}

# Mapping from IDs to names (for readability if needed)
id_to_name = {row["Input Id"]: row["Input Name"] for _, row in df.iterrows()}
id_to_name.update({row["Output Id"]: row["Output Name"] for _, row in df.iterrows()})

# Build parent and child maps
for _, row in df.iterrows():
    parent, child = row["Input Id"], row["Output Id"]
    parent_map.setdefault(child, []).append(parent)  
    child_map.setdefault(parent, []).append(child)  

# ⬅️ Function to recursively traverse upstream (parents)
def traverse_parents(node):
    for parent in parent_map.get(node, []):
        parent_name = id_to_name.get(parent, "").strip()

        # Always add the edge
        merged_edges.add((parent, node))
        unique_nodes.update([parent, node])

        # Recurse only if parent matches WK pattern
        if re.search(r'WK_\d{2}$|WK\d{2}$', parent_name):
            traverse_parents(parent)


# ⬇️ Function to recursively traverse downstream (children)
def traverse_children(node):
    if node in visited:
        return
    visited.add(node)

    if node in child_map:
        for child in child_map[node]:
            merged_edges.add((node, child))
            unique_nodes.update([node, child])
            descendants.add(child)
            traverse_children(child)

# 🚀 Process each target table
for target_table in target_tables:
    print(f"Processing Target Table for 2nd Diagram(WK): {target_table}...")

    # Step 1 (updated): Recursively collect parent edges (upstream)
    traverse_parents(target_table)

    # Step 2: Traverse Children (Downstream)
    visited = set()           # reset visited for each target
    descendants = set()

    if target_table in child_map:
        for direct_child in child_map[target_table]:
            merged_edges.add((target_table, direct_child))
            unique_nodes.update([target_table, direct_child])
            descendants.add(direct_child)
            traverse_children(direct_child)


graph = Graph2D()
parent_map_new = {}  
child_map_new = {}  

for parent, child in merged_edges:
    
    if pd.notna(child) and child != "":
        graph.add_node(parent)
        graph.add_node(child)
        graph.add_edge(parent, child)
        parent_map_new.setdefault(child, []).append(parent) 
        child_map_new.setdefault(parent, []).append(child)  
merged_edges_df = pd.DataFrame(list(merged_edges), columns=["Parent", "Child"])


merged_edges_df["Parent Name"] = merged_edges_df["Parent"].map(id_to_name)
merged_edges_df["Child Name"] = merged_edges_df["Child"].map(id_to_name)

# ✅ **Save to Excel**

### Change 4:In the line file_path change it to = r"Your local path\Final_Unique_Edges.xlsx"
file_path = edges_input_file
input_directory = os.path.dirname(file_path)

output_file_path = os.path.join(input_directory, "Merged_Edges.xlsx")
merged_edges_df.to_excel(output_file_path, index=False)




graph.place_nodes()
positions = graph.get_positions()







import pandas as pd # type: ignore
import os
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font, Alignment # type: ignore

import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def get_direct_relations_for_all_nodes(child_map, parent_map, id_to_name, output_file):
    result_data = []
    table_number = 1

    # 🔍 Gather all unique nodes from both maps
    all_nodes = set(child_map.keys()) | set(parent_map.keys())

    for node in sorted(all_nodes):
        node_name = id_to_name.get(node, str(node))

        # ✅ Get direct parents and children
        direct_parents = parent_map.get(node, [])
        direct_children = child_map.get(node, [])

        # Convert to strings & names
        parent_ids = [str(pid) for pid in direct_parents if pd.notna(pid)]
        parent_names = [id_to_name.get(pid, "Unknown") for pid in parent_ids]
        child_ids = [str(cid) for cid in direct_children if pd.notna(cid)]
        child_names = [id_to_name.get(cid, "Unknown") for cid in child_ids]

        # If there are no parents or children, still show one row with blanks
        max_rows = max(len(parent_ids), len(child_ids), 1)

        for i in range(max_rows):
            result_data.append({
                "Table Number": str(table_number) if i == 0 else "",
                "Target Table ID": str(node) if i == 0 else "",
                "Target Table Name": node_name if i == 0 else "",
                "One_Level_Parent IDs": parent_ids[i] if i < len(parent_ids) else "",
                "One_Level_Parent Names": parent_names[i] if i < len(parent_names) else "",
                "All_Descendant IDs": child_ids[i] if i < len(child_ids) else "",
                "All_Descendant Names": child_names[i] if i < len(child_names) else "",
            })

        # Add 2 blank spacer rows
        for _ in range(2):
            result_data.append({
                "Target Table ID": "",
                "Target Table Name": "",
                "One_Level_Parent IDs": "",
                "One_Level_Parent Names": "",
                "All_Descendant IDs": "",
                "All_Descendant Names": "",
            })

        table_number += 1

    # 📝 Export to Excel
    if result_data:
        result_df = pd.DataFrame(result_data)
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        if os.path.exists(output_file):
            os.remove(output_file)

        with pd.ExcelWriter(output_file, mode="w", engine="openpyxl") as writer:
            result_df.to_excel(writer, sheet_name="Table_Relations_WK", index=False)

        # Styling with openpyxl
        wb = load_workbook(output_file)
        ws = wb.active

        # Bold + center header
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # Center + wrap other cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(output_file)



### Change 5:In the line output_file_new change it to = r"Your local path\Final_Output.xlsx"
table_name = "Intermediate_2"  # or however you derive it
output_file = os.path.join(output_folder_path, f"{table_name}.xlsx")
output_file_new_2 = output_file



import pandas as pd

# Step 1: Load the Excel file and parse the relevant sheet with correct header row
file_path = input_file_path # adjust path if needed
xls = pd.ExcelFile(file_path)
df = xls.parse("バッチ一覧", header=2)

# Step 2: Filter where Target is '○'
df_target = df[df['Unnamed: 0'] == '○']

# Step 3: Get unique 'バッチ処理物理名'
unique_batch_names = df_target['バッチ処理物理名'].dropna().unique()

# Step 4: Filter original data for those batch names
df_filtered_batches = df[df['バッチ処理物理名'].isin(unique_batch_names)]

# Step 5: Filter rows where 'マッピング処理物理名' ends with 'ADD_MASTER'
df_final = df_filtered_batches[
    df_filtered_batches['マッピング処理物理名'].astype(str).str.endswith('ADD_MASTER')
]

# Step 6: Build dependency mapping
dependency_rows = []

for _, row in df_final.iterrows():
    input_id = row.get('テーブル/ビューID.1')
    input_name = row.get('テーブル名/ビュー名.1')

    # These may contain multiple entries separated by newline
    dep_ids_raw = str(row.get('テーブル/ビューID'))
    dep_names_raw = str(row.get('テーブル名/ビュー名'))

    # Split by newline and clean
    dep_ids = [x.strip() for x in dep_ids_raw.split('\n') if x.strip()]
    dep_names = [x.strip() for x in dep_names_raw.split('\n') if x.strip()]

    # Match each TBMT* id and pair it with its name
    for dep_id, dep_name in zip(dep_ids, dep_names):
        if dep_id.startswith("TBMT"):
            dependency_rows.append({
                "Input ID": input_id,
                "Input Name": input_name,
                "Dependency ID": dep_id,
                "Dependency Name": dep_name
            })

# Step 7: Convert to DataFrame
df_dependencies = pd.DataFrame(dependency_rows)
input_directory = os.path.dirname(file_path)
output_file_path = os.path.join(input_directory, "TBMT_Boxes.xlsx")
df_dependencies.to_excel(output_file_path, index=False)
# Optional: Save result to Excel
# df_dependencies.to_excel("input_dependency_mapping.xlsx", index=False)





get_direct_relations_for_all_nodes(child_map_new, parent_map_new, id_to_name, output_file_new_2)




excel = win32com.client.DispatchEx("Excel.Application")

wb = excel.Workbooks.Open(output_file_new_2)
ws = wb.Sheets.Add(After=wb.Sheets(1))  
ws.Name = "Diagram_WK"
shapes = ws.Shapes
table_shapes = {}


for node, (x, y) in positions.items():
    display_name = id_to_name.get(node, node)  
    shape = shapes.AddShape(13, x, y, 160, 60)
    shape.TextFrame.Characters().Text = display_name  
    shape.Fill.ForeColor.RGB = 0xEEDD99  
    shape.TextFrame.HorizontalAlignment = 3  
    shape.Line.ForeColor.RGB = 0x000000  
    shape.Line.Weight = 1 
    shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0x000000 
    shape.TextFrame2.TextRange.ParagraphFormat.Alignment = 2  
    shape.TextFrame2.VerticalAnchor = 3 
    table_shapes[node] = shape




for child, parents in graph.edges.items():
    for parent in parents:
        if parent in table_shapes and child in table_shapes:
            shape1 = table_shapes[parent]
            shape2 = table_shapes[child]

            connector = shapes.AddConnector(2, 0, 0, 0, 0)  

            if shape1.Left > shape2.Left and shape1.Top > shape2.Top:  

              connector.ConnectorFormat.BeginConnect(shape1, 3)  
              connector.ConnectorFormat.EndConnect(shape2, 5)  

            elif shape1.Left > shape2.Left and shape1.Top < shape2.Top:  
 
              connector.ConnectorFormat.BeginConnect(shape1, 3)  
              connector.ConnectorFormat.EndConnect(shape2, 5)  

            elif shape1.Left < shape2.Left and shape1.Top > shape2.Top:  

             connector.ConnectorFormat.BeginConnect(shape1, 5)  
             connector.ConnectorFormat.EndConnect(shape2, 4)  

            elif shape1.Left < shape2.Left and shape1.Top < shape2.Top:  

             connector.ConnectorFormat.BeginConnect(shape1, 4)  
             connector.ConnectorFormat.EndConnect(shape2, 3)  

            elif shape1.Left == shape2.Left:  

             connector.ConnectorFormat.BeginConnect(shape1, 3)  
             connector.ConnectorFormat.EndConnect(shape2, 3)  

            elif shape1.Top == shape2.Top:  

                 horizontal_distance = abs(shape1.Left - shape2.Left)

                 if horizontal_distance > 500:
        # Connect from side-to-side (port 2 = left-middle)
                    connector.ConnectorFormat.BeginConnect(shape1, 2)
                    connector.ConnectorFormat.EndConnect(shape2, 2)
                 else:
        # Default: connect from bottom-middle of shape1 to top-middle of shape2
                    connector.ConnectorFormat.BeginConnect(shape1, 5)
                    connector.ConnectorFormat.EndConnect(shape2, 3)
                


            else:  

              connector.ConnectorFormat.BeginConnect(shape1, 1)  
              connector.ConnectorFormat.EndConnect(shape2, 3)  
  

            connector.Line.ForeColor.RGB = 0x000000  
            connector.Line.Weight = 2  
            connector.Line.EndArrowheadStyle = 3  





  
for sink_node in graph.sink_nodes:
    sink_shape = table_shapes.get(sink_node)  
    if sink_shape:
        sink_shape.Fill.ForeColor.RGB = 0xC1E1C1   
        sink_shape.Line.ForeColor.RGB = 0x000000  
        sink_shape.Line.Weight = 1 
        sink_shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0x000000 
        sink_shape.TextFrame2.TextRange.Font.Bold = True 




for target_table in target_tables:
    target_shape = table_shapes.get(target_table)
    if target_shape:
        target_shape.Fill.ForeColor.RGB = 0x0000FF  
        target_shape.Line.ForeColor.RGB = 0x000000  
        target_shape.Line.Weight = 1  
        target_shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0xFFFFFF  
        target_shape.TextFrame2.TextRange.Font.Bold = True  



ws.Application.ActiveWindow.DisplayGridlines = False 
ws.Application.ActiveWindow.DisplayHeadings = False   
ws.Cells.Interior.ColorIndex = 2

import pandas as pd

# Load the summary input Excel file
summary_input_file = TBMT_Input_path  # 🔁 Update path
summary_df = pd.read_excel(summary_input_file)

# Group rows by Input ID & Name
grouped = summary_df.groupby(['Input ID', 'Input Name'])

# Convert to the same structure as before
custom_summary_boxes = []
for (input_id, input_name), group in grouped:
    dependencies = list(zip(group['Dependency ID'], group['Dependency Name']))
    custom_summary_boxes.append({
        "id": input_id,
        "name": input_name,
        "dependencies": dependencies
    })


# === CONFIG ===
box_width = 250
line_height = 22
x_spacing = 300  # Space between each box

for i, box in enumerate(custom_summary_boxes):
    node_id = box["id"]
    node_name = box["name"]
    deps = box["dependencies"]

    node_pos = positions.get(node_id)
    if not node_pos:
        print(f"⚠️ Node {node_id} not found in layout. Skipping box.")
        continue

    node_x, node_y = node_pos
    base_x = node_x + x_spacing  # Position to the right of the node
    base_y = node_y

    lines = [f"{node_name} ({node_id})"]
    lines.append("--------------------")
    lines += [f"{name} ({id_})" for id_, name in deps]
    text_content = "\n".join(lines)

    shape = ws.Shapes.AddShape(1, base_x, base_y, box_width, line_height * len(lines)+1)
    shape.TextFrame.Characters().Text = text_content

    # Style the box
    shape.Fill.ForeColor.RGB = 0xDDEEFF
    shape.Line.ForeColor.RGB = 0x000000
    shape.Line.Weight = 1.5
    shape.TextFrame2.TextRange.Font.Size = 10
    shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 0x000000
    shape.TextFrame2.TextRange.ParagraphFormat.Alignment = 1  # Left-align
    shape.TextFrame2.VerticalAnchor = 1  # Top-align
    shape.TextFrame.AutoSize = True
    # ✅ After creating the shape (summary box)
    original_shape = table_shapes.get(node_id)

    if original_shape:
        connector = ws.Shapes.AddConnector(2, 0, 0, 0, 0)  # Straight line

    # Connect original shape's right side ➝ to box's left side
        connector.ConnectorFormat.BeginConnect(original_shape, 5)  # Right of node
        connector.ConnectorFormat.EndConnect(shape, 2)             # Left of box

        connector.Line.ForeColor.RGB = 0x666666  # Gray line (subtle)
        connector.Line.Weight = 1.25
        connector.Line.EndArrowheadStyle = 2  # Small arrow
    else:
        print(f"⚠️ Could not find original shape for {node_id} to connect.")

    



wb.Save()
excel.Quit()




# print(f"Output saved as Final_Output_2.xlsx ")



import openpyxl
import os



  # or however you derive it
output_file = os.path.join(output_folder_path, "Final_Output.xlsx")
output_file_new = output_file
import os
import win32com.client

import os
import win32com.client

def merge_excel_files(file1_path, file2_path, output_path):
    if os.path.exists(output_path):
        os.remove(output_path)

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        # Create a new workbook (has 1 default sheet)
        wb_output = excel.Workbooks.Add()
        default_sheet = wb_output.Sheets(1)  # Keep reference to delete later

        # Open both source workbooks
        wb1 = excel.Workbooks.Open(file1_path)
        wb2 = excel.Workbooks.Open(file2_path)

        def copy_all_sheets(src_wb):
            for sheet in src_wb.Sheets:
                sheet.Copy(After=wb_output.Sheets(wb_output.Sheets.Count))

        # Copy from both files
        copy_all_sheets(wb1)
        copy_all_sheets(wb2)

        # Now it's safe to delete the default sheet
        if wb_output.Sheets.Count > 1:
            default_sheet.Delete()

        # Save merged workbook
        wb_output.SaveAs(output_path)
        print(f"Files merged successfully into: Final_Output.xlsx")

        # Cleanup
        wb1.Close(False)
        wb2.Close(False)
        wb_output.Close(False)

    except Exception as e:
        print(f"❌ Error during merge: {e}")

    finally:
        excel.Quit()









import time

def wait_until_file_is_ready(filepath, timeout=30):
    """Wait until a file exists and can be opened (not locked), up to timeout seconds."""
    start_time = time.time()
    while True:
        if os.path.exists(filepath):
            try:
                with open(filepath, 'rb'):
                    return True  # File is ready
            except PermissionError:
                pass  # File exists but still in use

        if time.time() - start_time > timeout:
            raise TimeoutError(f"File '{filepath}' is not ready after {timeout} seconds.")
        time.sleep(1)  # Wait 1 second before checking again

wait_until_file_is_ready(output_file_new_2)

merge_excel_files(
    file1_path=output_file_new_1,
    file2_path=output_file_new_2,
    output_path=output_file_new
)










































